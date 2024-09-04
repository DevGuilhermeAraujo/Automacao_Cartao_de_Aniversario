import os
import sys
import openpyxl
from PIL import Image, ImageDraw, ImageFont
import tkinter as tk
from tkinter import filedialog, messagebox

def get_resource_path(relative_path):
    """Obtém o caminho absoluto para o recurso, seja na execução local ou em um executável."""
    try:
        # PyInstaller cria uma pasta temporária e armazena o caminho nela
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

def selecionar_arquivo():
    """Abre uma janela para selecionar o arquivo da planilha Excel."""
    root = tk.Tk()
    root.withdraw()  # Oculta a janela principal do Tkinter

    caminho_arquivo = filedialog.askopenfilename(
        title="Selecione a planilha do Excel",
        filetypes=[("Planilhas do Excel", "*.xlsx *.xls")]
    )

    if not caminho_arquivo:
        messagebox.showerror("Erro", "Nenhum arquivo foi selecionado.")
        sys.exit()  # Encerra o programa se nenhum arquivo for selecionado

    return caminho_arquivo

def obter_parametros_planilha(planilha):
    # Conta o número de pessoas na planilha
    num_pessoas = len(list(planilha.iter_rows(min_row=2)))

    # Define o tamanho da fonte e outros parâmetros com base no número de pessoas
    if num_pessoas <= 9:
        font_data_size = 200
        font_nome_size = 200
        incrementoY = 250
        coordenadaX = [1600, 1800, 1880, 2300]
        linha_vertical_x = 2200
    elif 9 < num_pessoas <= 11:
        font_data_size = 170
        font_nome_size = 170
        incrementoY = 220
        coordenadaX = [1600, 1780, 1855, 2300]
        linha_vertical_x = 2200
    elif 11 < num_pessoas <= 13:
        font_data_size = 140
        font_nome_size = 140
        incrementoY = 190
        coordenadaX = [1600, 1750, 1800, 2200]
        linha_vertical_x = 2100
    elif 13 < num_pessoas <= 15:
        font_data_size = 110
        font_nome_size = 110
        incrementoY = 160
        coordenadaX = [1600, 1720, 1770, 1970]
        linha_vertical_x = 1920
    else:
        font_data_size = 90
        font_nome_size = 90
        incrementoY = 130
        coordenadaX = [1600, 1700, 1750, 1950]
        linha_vertical_x = 1900

    # Retorna os parâmetros calculados
    return {
        'num_pessoas': num_pessoas,
        'font_data_size': font_data_size,
        'font_nome_size': font_nome_size,
        'incrementoY': incrementoY,
        'coordenadaX': coordenadaX,
        'linha_vertical_x': linha_vertical_x
    }

def gerar_cartao(planilha, parametros, imagem_base, titulo, subtitulo, arquivo_saida):
    # Carrega a imagem base
    image = Image.open(get_resource_path(imagem_base))
    desenhar = ImageDraw.Draw(image)
    
    # Carrega as fontes
    font_titulo = ImageFont.truetype(get_resource_path('./Scansky Condensed Semi Bold.ttf'), 350)
    font_data = ImageFont.truetype(get_resource_path('./Scansky Condensed Semi Bold.ttf'), parametros['font_data_size'])
    font_nome = ImageFont.truetype(get_resource_path('./Scansky Condensed Semi Bold.ttf'), parametros['font_nome_size'])

    # Desenha os títulos
    desenhar.text((1600, 650), titulo, fill='black', font=font_titulo)
    desenhar.text((1600, 1000), subtitulo, fill='black', font=font_titulo)

    # Define a coordenada Y inicial
    coordenadaY = 1550

    # Itera pelas linhas da planilha, começando da segunda linha
    for linha in planilha.iter_rows(min_row=2):
        data = linha[0].value
        nome = linha[1].value
        
        # Ignora linhas onde data ou nome sejam None ou vazios
        if data and nome:
            # Separa o dia e o mês da data
            dia = data[:2]
            mes = data[3:]

            # Desenha a data e o nome
            desenhar.text((parametros['coordenadaX'][0], coordenadaY), dia, fill='black', font=font_data)
            desenhar.text((parametros['coordenadaX'][1], coordenadaY), "/", fill='black', font=font_data)
            desenhar.text((parametros['coordenadaX'][2], coordenadaY), mes, fill='black', font=font_data)
            
            # Desenha a linha vertical
            linha_inicio = (parametros['linha_vertical_x'], coordenadaY)
            linha_fim = (parametros['linha_vertical_x'], coordenadaY + parametros['incrementoY'])
            desenhar.line([linha_inicio, linha_fim], fill='black', width=10)
            
            # Desenha o nome
            desenhar.text((parametros['coordenadaX'][3], coordenadaY), nome, fill='black', font=font_nome)
            
            # Incrementa a coordenada Y
            coordenadaY += parametros['incrementoY']

    # Salva a imagem final
    image.save(arquivo_saida)

def main():
    # Seleciona o arquivo de planilha
    caminho_planilha = selecionar_arquivo()

    # Carrega a planilha do Excel
    workbook_funcionarios = openpyxl.load_workbook(caminho_planilha)
    planilha_aniversario = workbook_funcionarios['aniversario']
    planilha_tempoDeEmpresa = workbook_funcionarios['tempoEmpresa']

    # Obter os parâmetros para cada planilha
    parametros_aniversario = obter_parametros_planilha(planilha_aniversario)
    parametros_tempoDeEmpresa = obter_parametros_planilha(planilha_tempoDeEmpresa)

    # Meses mapeados para os títulos
    mes_aniversario = planilha_aniversario['A2'].value[3:5]
    mes_tempoDeEmpresa = planilha_tempoDeEmpresa['A2'].value[3:5]

    meses = {
        "01": "JANEIRO",
        "02": "FEVEREIRO",
        "03": "MARÇO",
        "04": "ABRIL",
        "05": "MAIO",
        "06": "JUNHO",
        "07": "JULHO",
        "08": "AGOSTO",
        "09": "SETEMBRO",
        "10": "OUTUBRO",
        "11": "NOVEMBRO",
        "12": "DEZEMBRO"
    }

    mes_extenso_aniversario = meses.get(mes_aniversario, "Mês desconhecido")
    mes_extenso_tempoDeEmpresa = meses.get(mes_tempoDeEmpresa, "Mês desconhecido")

    # Gerar cartão de aniversário
    gerar_cartao(
        planilha_aniversario, 
        parametros_aniversario,
        './modelo(colaborador).png', 
        "ANIVERSARIANTES", 
        f"DO MÊS DE {mes_extenso_aniversario}",
        'cartao_aniversario.png'
    )

    # Gerar cartão de tempo de empresa
    gerar_cartao(
        planilha_tempoDeEmpresa, 
        parametros_tempoDeEmpresa,
        './modelo(tempoDeEmpresa).png', 
        "ANIVERSARIANTES", 
        f"POR TEMPO DE EMPRESA/{mes_extenso_tempoDeEmpresa}",
        'cartao_tempoDeEmpresa.png'
    )

    messagebox.showinfo("Sucesso", "Os cartões foram gerados com sucesso!")

if __name__ == "__main__":
    main()
